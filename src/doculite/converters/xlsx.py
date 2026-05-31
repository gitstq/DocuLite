"""
📊 DocuLite XLSX Converter Module

Converter for Excel spreadsheets.
"""

from pathlib import Path
from typing import Optional, Dict, Any, Union, List
import io
import re

from .base import BaseConverter
from ..types import ConversionResult, DocumentType
from ..exceptions import ConversionError
from ..utils import clean_text


class XlsxConverter(BaseConverter):
    """Converter for Excel spreadsheets (.xlsx, .xls)."""
    
    SUPPORTED_TYPES = [DocumentType.XLSX, DocumentType.XLS]
    
    def __init__(self, config: Optional[Dict[str, Any]] = None):
        super().__init__(config)
        self.max_rows_per_sheet = config.get('max_rows_per_sheet', 1000)
        self.include_formulas = config.get('include_formulas', False)
        
    def _get_openpyxl(self):
        """Lazy load openpyxl."""
        try:
            import openpyxl
            return openpyxl
        except ImportError:
            raise ConversionError(
                "openpyxl is required for Excel conversion. "
                "Install with: pip install 'doculite[xlsx]'"
            )
    
    def convert(self, file_path: Union[str, Path, io.BytesIO], **kwargs) -> ConversionResult:
        """
        Convert an Excel spreadsheet to markdown.
        
        Args:
            file_path: Path to the Excel file or file-like object
            **kwargs: Additional options
            
        Returns:
            ConversionResult with markdown content
        """
        openpyxl = self._get_openpyxl()
        
        try:
            if isinstance(file_path, (str, Path)):
                path = self.validate_file(file_path)
                wb = openpyxl.load_workbook(str(path), data_only=not self.include_formulas)
            else:
                file_path.seek(0)
                wb = openpyxl.load_workbook(file_path, data_only=not self.include_formulas)
            
            markdown_parts = []
            metadata = {}
            pages_info = []
            tables_info = []
            
            # Extract properties
            props = wb.properties
            metadata = {
                "title": props.title or "",
                "creator": props.creator or "",
                "subject": props.subject or "",
                "description": props.description or "",
                "created": str(props.created) if props.created else "",
                "modified": str(props.modified) if props.modified else "",
                "sheet_count": len(wb.sheetnames),
                "sheets": wb.sheetnames,
            }
            
            # Process each sheet
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                
                sheet_md = self._convert_sheet(sheet, sheet_name)
                if sheet_md:
                    markdown_parts.append(sheet_md)
                    
                    tables_info.append({
                        "sheet_name": sheet_name,
                        "rows": sheet.max_row,
                        "columns": sheet.max_column,
                    })
            
            wb.close()
            
            # Combine all parts
            markdown = "\n\n".join(markdown_parts)
            markdown = self.post_process_markdown(markdown)
            
            # Generate plain text
            text = re.sub(r'[#*_`\[\]|]', '', markdown)
            text = clean_text(text)
            
            return ConversionResult(
                markdown=markdown,
                text=text,
                metadata=metadata,
                pages=pages_info,
                images=[],
                tables=tables_info,
                links=[],
            )
            
        except Exception as e:
            raise ConversionError(
                f"Failed to convert Excel file: {str(e)}",
                {"file_path": str(file_path) if isinstance(file_path, (str, Path)) else "stream"}
            )
    
    def _convert_sheet(self, sheet, sheet_name: str) -> str:
        """Convert a worksheet to markdown."""
        lines = []
        
        # Add sheet header
        lines.append(f"## Sheet: {sheet_name}")
        lines.append("")
        
        # Check if sheet has data
        if sheet.max_row == 0 or sheet.max_column == 0:
            lines.append("*Empty sheet*")
            return '\n'.join(lines)
        
        # Limit rows to prevent huge outputs
        max_row = min(sheet.max_row, self.max_rows_per_sheet)
        
        # Build table
        rows = []
        for row_idx in range(1, max_row + 1):
            row_data = []
            for col_idx in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                value = cell.value
                
                # Format value
                if value is None:
                    value = ""
                elif isinstance(value, (int, float)):
                    value = str(value)
                else:
                    value = str(value).strip()
                    # Truncate long cells
                    if len(value) > 100:
                        value = value[:97] + "..."
                
                row_data.append(value)
            
            # Skip empty rows
            if any(cell for cell in row_data):
                rows.append(row_data)
        
        if not rows:
            lines.append("*No data*")
            return '\n'.join(lines)
        
        # Determine if first row is a header
        # Heuristic: first row has different data types or all strings
        is_header = self._is_likely_header(rows)
        
        # Build markdown table
        if is_header and len(rows) > 1:
            header = rows[0]
            data_rows = rows[1:]
            
            lines.append("| " + " | ".join(header) + " |")
            lines.append("| " + " | ".join(["---"] * len(header)) + " |")
            
            for row in data_rows:
                # Pad row to match header length
                while len(row) < len(header):
                    row.append("")
                row = row[:len(header)]
                lines.append("| " + " | ".join(row) + " |")
        else:
            # No header, just data
            max_cols = max(len(row) for row in rows)
            for row in rows:
                while len(row) < max_cols:
                    row.append("")
                lines.append("| " + " | ".join(row) + " |")
        
        # Add note if truncated
        if sheet.max_row > self.max_rows_per_sheet:
            lines.append("")
            lines.append(f"*Note: Showing {self.max_rows_per_sheet} of {sheet.max_row} rows*")
        
        return '\n'.join(lines)
    
    def _is_likely_header(self, rows: List[List[str]]) -> bool:
        """Heuristic to determine if first row is a header."""
        if len(rows) < 2:
            return False
        
        first_row = rows[0]
        second_row = rows[1]
        
        # Check if first row is all strings and second row has numbers
        first_all_strings = all(isinstance(cell, str) for cell in first_row if cell)
        second_has_numbers = any(self._is_number(cell) for cell in second_row if cell)
        
        return first_all_strings and second_has_numbers
    
    def _is_number(self, value: str) -> bool:
        """Check if string represents a number."""
        try:
            float(value.replace(',', ''))
            return True
        except (ValueError, AttributeError):
            return False
